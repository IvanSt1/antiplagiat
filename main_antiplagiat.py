import ast
import difflib
import os
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def get_ast_structure(file_path):
    """Парсит Python файл и возвращает упрощённую структуру AST."""
    with open(file_path, 'r') as f:
        tree = ast.parse(f.read())
    
    structure = []
    for node in ast.walk(tree):
        if isinstance(node, ast.FunctionDef):
            structure.append("FunctionDef")
        elif isinstance(node, ast.ClassDef):
            structure.append("ClassDef")
        elif isinstance(node, ast.If):
            structure.append("If")
        elif isinstance(node, ast.For):
            structure.append("For")
        elif isinstance(node, ast.While):
            structure.append("While")
        elif isinstance(node, ast.Try):
            structure.append("Try")
        elif isinstance(node, ast.With):
            structure.append("With")
        elif isinstance(node, ast.Return):
            structure.append("Return")
        elif isinstance(node, ast.Assign):
            structure.append("Assign")
        elif isinstance(node, ast.Expr):
            structure.append("Expr")
        elif isinstance(node, ast.Call):
            structure.append("Call")
    return structure

def compare_ast_structures(file1, file2):
    """Сравнивает структуры AST двух Python файлов и возвращает коэффициент схожести в процентах."""
    structure1 = get_ast_structure(file1)
    structure2 = get_ast_structure(file2)
    similarity_ratio = difflib.SequenceMatcher(None, structure1, structure2).ratio()
    return similarity_ratio * 100  # Конвертируем в проценты

def check_plagiarism_across_all_assignments(base_path, similarity_threshold=90):
    """Проходит по всем заданиям всех классов и создает CSV и XLSX-файлы для каждого задания с отметкой 100% схожести."""
    output_folder = "plagiarism_reports"
    os.makedirs(output_folder, exist_ok=True)

    all_solutions = {}
    students_with_100 = set()
    summary_report = {}

    for class_dir in os.listdir(base_path):
        class_path = os.path.join(base_path, class_dir)
        if not os.path.isdir(class_path):
            continue
        
        for student_dir in os.listdir(class_path):
            student_path = os.path.join(class_path, student_dir)
            if not os.path.isdir(student_path):
                continue
            
            for assignment_file in os.listdir(student_path):
                assignment_path = os.path.join(student_path, assignment_file)
                assignment_name = os.path.splitext(assignment_file)[0]
                
                student_name = f"{class_dir} {student_dir}"
                
                if assignment_name not in all_solutions:
                    all_solutions[assignment_name] = []
                
                all_solutions[assignment_name].append((student_name, assignment_path))
    
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for assignment_name, solutions in all_solutions.items():
        output_csv = os.path.join(output_folder, f"{assignment_name}_plagiarism.csv")
        output_xlsx = os.path.join(output_folder, f"{assignment_name}_plagiarism.xlsx")
        
        students = [student for student, _ in solutions]
        results = {student: {other_student: "" for other_student in students} for student in students}
        
        for i, (student1, file1) in enumerate(solutions):
            for j, (student2, file2) in enumerate(solutions):
                if i < j:
                    similarity = compare_ast_structures(file1, file2)
                    if similarity == 100.0:
                        students_with_100.update([student1, student2])
                    results[student1][student2] = f"{similarity:.2f}"
                    results[student2][student1] = f"{similarity:.2f}"
                        # Добавляем совпадение в сводный отчет
                    if similarity >= similarity_threshold:
                        if student1 not in summary_report:
                            summary_report[student1] = {}
                        if student2 not in summary_report:
                            summary_report[student2] = {}
                        summary_report[student1].setdefault(assignment_name, []).append(student2)
                        summary_report[student2].setdefault(assignment_name, []).append(student1)
        
        # Запись CSV-файла для задания
        with open(output_csv, mode="w", newline='', encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            header = [""] + students
            writer.writerow(header)
            for student, comparisons in results.items():
                row = [student] + [comparisons[other_student] for other_student in students]
                writer.writerow(row)
        
        # Запись XLSX-файла для задания
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = assignment_name
        sheet.append([""] + students)
        for student, comparisons in results.items():
            row = [student] + [comparisons[other_student] for other_student in students]
            sheet.append(row)
        
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=len(students) + 1, max_col=len(students) + 1):
            for cell in row:
                if cell.value == "100.00":
                    cell.fill = red_fill
        
        workbook.save(output_xlsx)
        print(f"Отчет для задания {assignment_name} сохранен как {output_csv} и {output_xlsx}")

    # Создание сводного отчета
    summary_xlsx = os.path.join(output_folder, "summary_report.xlsx")
    summary_workbook = Workbook()
    summary_sheet = summary_workbook.active
    summary_sheet.title = "Summary Report"
    assignments = sorted(all_solutions.keys())
    summary_sheet.append(["Ученик"] + assignments)

    for student in sorted(summary_report.keys()):
        row = [student]
        for assignment in assignments:
            matches = summary_report[student].get(assignment, [])
            row.append(", ".join(matches) if matches else "")
        summary_sheet.append(row)

    summary_workbook.save(summary_xlsx)
    print(f"Сводный отчет сохранен как {summary_xlsx}")

    # Сохранение списка учеников с совпадением 100% в текстовый файл
    if students_with_100:
        with open("students_with_100.txt", "w", encoding="utf-8") as txtfile:
            for student in sorted(students_with_100):
                txtfile.write(student + "\n")
        print("Список учеников с совпадением 100% сохранен в students_with_100.txt")

# Пример использования
base_path = "solutions"
check_plagiarism_across_all_assignments(base_path)
